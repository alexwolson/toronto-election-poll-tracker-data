#!/usr/bin/env python3
"""Fetch official Toronto councillor results for ward-model calibration.

Toronto used 44 wards from 2003 through 2014 and 25 wards from 2018 onward.
The stable-boundary transitions 2003→2006, 2006→2010, 2010→2014, and
2018→2022 provide historical incumbent re-election cases without pretending
that a candidate crossing a boundary change represents the same ward contest.

Run: uv run scripts/fetch_historical_council_results.py
"""

from __future__ import annotations

import io
import re
import sys
import unicodedata
import zipfile
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd
import xlrd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.fetch_elections import (
    GENERAL_RESULTS_PACKAGE,
    download_bytes,
    fetch_resources,
)

OUTPUT_PATH = Path("data/raw/elections/historical_council_results.csv")
YEARS = (2003, 2006, 2010, 2014, 2018, 2022)


def candidate_key(value: object) -> str:
    """Order-insensitive key for the City's changing first/last-name formats."""
    normalized = unicodedata.normalize("NFKD", str(value))
    ascii_name = "".join(char for char in normalized if not unicodedata.combining(char))
    tokens = re.findall(r"[a-z0-9]+", ascii_name.lower())
    return " ".join(sorted(tokens))


def _ward_number(sheet_name: str) -> int | None:
    match = re.search(r"(?:ward\s*)(\d+)", sheet_name, flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def _office_file(archive: zipfile.ZipFile) -> str:
    matches = [
        name
        for name in archive.namelist()
        if "councillor" in name.lower()
        and name.lower().endswith((".xls", ".xlsx"))
    ]
    if len(matches) != 1:
        raise ValueError(f"Expected one councillor workbook; found {matches}")
    return matches[0]


def _xlsx_rows(raw: bytes) -> Iterable[tuple[str, list[list[object]]]]:
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        yield sheet_name, [list(row) for row in worksheet.iter_rows(values_only=True)]


def _xls_rows(raw: bytes) -> Iterable[tuple[str, list[list[object]]]]:
    workbook = xlrd.open_workbook(file_contents=raw)
    for sheet_name in workbook.sheet_names():
        worksheet = workbook.sheet_by_name(sheet_name)
        yield sheet_name, [worksheet.row_values(index) for index in range(worksheet.nrows)]


def parse_workbook(raw: bytes, suffix: str, year: int, source_url: str) -> list[dict]:
    records: list[dict] = []
    sheets = _xlsx_rows(raw) if suffix == ".xlsx" else _xls_rows(raw)
    for sheet_name, rows in sheets:
        ward = _ward_number(sheet_name)
        if ward is None:
            continue

        header_index = next(
            (
                index
                for index, row in enumerate(rows)
                if row
                and str(row[0]).strip().lower() in {"name", "subdivision"}
                and any(str(value).strip().lower() == "total" for value in row)
            ),
            None,
        )
        if header_index is None:
            raise ValueError(f"{year} Ward {ward}: total header not found")
        header = rows[header_index]
        total_column = next(
            index
            for index, value in enumerate(header)
            if str(value).strip().lower() == "total"
        )

        ward_rows: list[dict] = []
        for row in rows[header_index + 1 :]:
            if not row or not str(row[0]).strip():
                continue
            candidate_name = str(row[0]).strip()
            if "total" in candidate_name.lower():
                continue
            if total_column >= len(row):
                continue
            votes = row[total_column]
            if not isinstance(votes, (int, float)):
                continue
            ward_rows.append(
                {
                    "election_year": year,
                    "ward": ward,
                    "boundary_era": "44_ward" if year <= 2014 else "25_ward",
                    "candidate_name": candidate_name,
                    "candidate_key": candidate_key(candidate_name),
                    "votes": int(votes),
                    "is_acclaimed": False,
                    "source_url": source_url,
                }
            )

        if not ward_rows:
            raise ValueError(f"{year} Ward {ward}: no candidate totals parsed")
        total_votes = sum(row["votes"] for row in ward_rows)
        if total_votes <= 0:
            if len(ward_rows) != 1:
                raise ValueError(f"{year} Ward {ward}: non-positive vote total")
            ward_rows[0]["vote_share"] = 1.0
            ward_rows[0]["is_winner"] = True
            ward_rows[0]["is_acclaimed"] = True
            records.extend(ward_rows)
            continue
        top_votes = max(row["votes"] for row in ward_rows)
        if sum(row["votes"] == top_votes for row in ward_rows) != 1:
            raise ValueError(f"{year} Ward {ward}: election winner is tied")
        for row in ward_rows:
            row["vote_share"] = row["votes"] / total_votes
            row["is_winner"] = row["votes"] == top_votes
        records.extend(ward_rows)
    return records


def validate_results(results: pd.DataFrame) -> None:
    expected_wards = {
        2003: 44,
        2006: 44,
        2010: 44,
        2014: 44,
        2018: 25,
        2022: 25,
    }
    if set(results["election_year"].unique()) != set(YEARS):
        raise ValueError("Historical councillor results do not cover every configured year")
    for year, expected in expected_wards.items():
        year_rows = results[results["election_year"] == year]
        if year_rows["ward"].nunique() != expected:
            raise ValueError(
                f"{year}: expected {expected} wards, got {year_rows['ward'].nunique()}"
            )
        if year_rows.duplicated(["ward", "candidate_key"]).any():
            raise ValueError(f"{year}: duplicate normalized candidate in a ward")
        winners = year_rows.groupby("ward")["is_winner"].sum()
        if not (winners == 1).all():
            raise ValueError(f"{year}: every ward must have exactly one winner")
        share_sums = year_rows.groupby("ward")["vote_share"].sum()
        if not ((share_sums - 1.0).abs() < 1e-9).all():
            raise ValueError(f"{year}: ward vote shares do not sum to one")


def main() -> None:
    resources = {
        int(match.group(1)): resource
        for resource in fetch_resources(GENERAL_RESULTS_PACKAGE)
        if (match := re.match(r"(\d{4})-results", resource.get("name", "")))
    }
    records: list[dict] = []
    for year in YEARS:
        resource = resources.get(year)
        if resource is None:
            raise RuntimeError(f"No official result archive found for {year}")
        print(f"Fetching {year} councillor results...")
        archive = zipfile.ZipFile(io.BytesIO(download_bytes(resource["url"])))
        workbook_name = _office_file(archive)
        workbook = archive.read(workbook_name)
        suffix = Path(workbook_name).suffix.lower()
        year_records = parse_workbook(workbook, suffix, year, resource["url"])
        records.extend(year_records)
        print(f"  Parsed {len(year_records)} candidate-ward rows")

    results = pd.DataFrame(records).sort_values(
        ["election_year", "ward", "votes"], ascending=[True, True, False]
    )
    validate_results(results)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    results.to_csv(OUTPUT_PATH, index=False)
    print(
        f"Written {OUTPUT_PATH}: {len(results)} candidate rows across "
        f"{results['election_year'].nunique()} elections"
    )


if __name__ == "__main__":
    main()
