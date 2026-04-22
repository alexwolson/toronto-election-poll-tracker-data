#!/usr/bin/env python3
"""Fetch Toronto municipal election results from Toronto Open Data.

Downloads XLSX files (via ZIP archives where needed), extracts ward-level
mayoral vote totals and registered-elector counts, and saves clean CSVs to
data/raw/elections/.

Outputs:
  data/raw/elections/mayoral_results.csv     -- ward x candidate x election
  data/raw/elections/registered_electors.csv -- ward x election x electors

Run: uv run scripts/fetch_elections.py
"""

from __future__ import annotations

import io
import json
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import requests

CKAN_BASE = "https://ckan0.cf.opendata.inter.prod-toronto.ca/api/3/action"

# General election results (ZIP of XLSX files)
GENERAL_RESULTS_PACKAGE = "election-results-official"

# By-election results (individual XLSX resources)
BY_ELECTION_RESULTS_PACKAGE = "elections-official-by-election-results"
BY_ELECTION_RESULTS_NAMES = {
    "2023 Office of the Mayor": 2023,
}

# Voter statistics (registered electors) — one XLSX per election
VOTER_STATS_PACKAGE = "elections-voter-statistics"
VOTER_STATS_NAMES = {
    "2022-voter-statistics": 2022,
    "2018-voter-statistics": 2018,
}

BY_ELECTION_VOTER_STATS_PACKAGE = "elections-by-election-voter-statistics"
BY_ELECTION_VOTER_STATS_NAMES = {
    "2023-mayoral-by-election-voter-statistics": 2023,
}

OUTPUT_DIR = Path("data/raw/elections")


# ---------------------------------------------------------------------------
# CKAN helpers
# ---------------------------------------------------------------------------


def fetch_resources(package_id: str) -> list[dict]:
    url = f"{CKAN_BASE}/package_show"
    r = requests.get(url, params={"id": package_id}, timeout=30)
    r.raise_for_status()
    data = r.json()
    if not data["success"]:
        raise RuntimeError(f"CKAN API error for {package_id}: {data.get('error')}")
    return data["result"]["resources"]


def download_bytes(url: str, timeout: int = 120) -> bytes:
    r = requests.get(url, timeout=timeout)
    r.raise_for_status()
    return r.content


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------


def _load_workbook_bytes(
    data: bytes, resolve_formulas: bool = False
) -> openpyxl.Workbook:
    # read_only=True is faster for large files but does not resolve cached formula values.
    # Set resolve_formulas=True for sheets that use cell-reference formulas (e.g. 2018
    # voter statistics) so Ward column values are integers, not formula strings.
    if resolve_formulas:
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def _extract_mayor_xlsx_from_zip(zip_bytes: bytes, year: int) -> bytes:
    """Return the bytes of the Mayor XLSX inside a general-election ZIP."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        mayor_files = [n for n in names if "Mayor" in n and n.endswith(".xlsx")]
        if not mayor_files:
            raise ValueError(f"{year}: no Mayor XLSX found in ZIP. Files: {names}")
        if len(mayor_files) > 1:
            raise ValueError(
                f"{year}: multiple Mayor XLSXs found in ZIP: {mayor_files}"
            )
        return zf.read(mayor_files[0])


def _parse_ward_totals(wb: openpyxl.Workbook, year: int) -> list[dict]:
    """Extract per-ward mayoral vote totals from a poll-by-poll XLSX workbook.

    The workbook has one sheet per ward named "Ward N".  Each sheet has:
      row 0: ward name header
      row 1: "Subdivision" | poll numbers ... | "Total"
      row 2: "Mayor" label
      rows 3+: candidate name | poll votes ... | total votes
      last row: "City Ward N Totals" | ... | grand total (votes cast)

    Returns a list of dicts with keys: year, ward, candidate, votes.
    """
    records = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Ward "):
            continue
        ward_num = int(sheet_name.split()[1])
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find Total column index from the header row (row index 1)
        header = rows[1]
        try:
            total_col = list(header).index("Total")
        except ValueError:
            raise ValueError(
                f"{year} Ward {ward_num}: no 'Total' column found in header row"
            )

        # Rows 3+ are candidates until the "City Ward N Totals" summary row
        for row in rows[3:]:
            candidate = row[0]
            if candidate is None:
                continue
            name = str(candidate).strip()
            if name.startswith("City Ward"):
                # This is the ward totals summary row — skip it
                continue
            votes_raw = row[total_col]
            if votes_raw is None or not isinstance(votes_raw, (int, float)):
                continue
            records.append(
                {
                    "year": year,
                    "ward": ward_num,
                    "candidate": name,
                    "votes": int(votes_raw),
                }
            )
    return records


def _parse_eligible_electors(wb: openpyxl.Workbook, year: int) -> list[dict]:
    """Extract per-ward total eligible electors from a voter statistics workbook.

    The workbook has a data sheet whose first row is a header. Relevant columns:
      'Ward', 'Total Eligible Electors'

    Returns a list of dicts with keys: year, ward, eligible_electors.
    """
    # Find the data sheet (not 'readme', 'notes', or LTC locations sheets)
    data_sheet = None
    for name in wb.sheetnames:
        lower = name.lower()
        if lower in ("readme", "notes") or "ltc" in lower or "read me" in lower:
            continue
        # Verify it has the expected header
        ws_candidate = wb[name]
        first_row = next(ws_candidate.iter_rows(max_row=1, values_only=True), ())
        if "Ward" in first_row:
            data_sheet = ws_candidate
            break
    if data_sheet is None:
        raise ValueError(f"{year} voter stats: could not identify data sheet")

    rows = list(data_sheet.iter_rows(values_only=True))
    header = list(rows[0])

    try:
        ward_col = header.index("Ward")
        eligible_col = header.index("Total Eligible Electors")
    except ValueError as e:
        raise ValueError(f"{year} voter stats: missing column: {e}")

    ward_totals: dict[int, int] = defaultdict(int)
    for row in rows[1:]:
        ward = row[ward_col]
        if not isinstance(ward, int):
            continue
        eligible = row[eligible_col]
        if eligible is not None and isinstance(eligible, (int, float)):
            ward_totals[ward] += int(eligible)

    return [
        {"year": year, "ward": ward, "eligible_electors": total}
        for ward, total in sorted(ward_totals.items())
    ]


# ---------------------------------------------------------------------------
# Main fetch logic
# ---------------------------------------------------------------------------


def fetch_mayoral_results() -> list[dict]:
    """Fetch and parse ward-level mayoral vote totals for all available elections."""
    all_records: list[dict] = []

    # --- General elections (ZIP → Mayor XLSX) ---
    print("Fetching general election results...")
    resources = fetch_resources(GENERAL_RESULTS_PACKAGE)
    for resource in resources:
        name = resource.get("name", "")
        fmt = resource.get("format", "").upper()
        # Identify year from resource name, e.g. "2022-results" or "2018-results"
        year = None
        for yr in (2022, 2018):
            if str(yr) in name:
                year = yr
                break
        if year is None or fmt != "ZIP":
            continue
        print(f"  Downloading {name} ({year})...")
        zip_bytes = download_bytes(resource["url"])
        xlsx_bytes = _extract_mayor_xlsx_from_zip(zip_bytes, year)
        wb = _load_workbook_bytes(xlsx_bytes)
        records = _parse_ward_totals(wb, year)
        all_records.extend(records)
        print(f"    -> {len(records)} candidate-ward rows")

    # --- 2023 mayoral by-election (standalone XLSX) ---
    print("Fetching 2023 mayoral by-election results...")
    by_election_resources = fetch_resources(BY_ELECTION_RESULTS_PACKAGE)
    for resource in by_election_resources:
        name = resource.get("name", "")
        year = BY_ELECTION_RESULTS_NAMES.get(name)
        if year is None:
            continue
        print(f"  Downloading {name}...")
        xlsx_bytes = download_bytes(resource["url"])
        wb = _load_workbook_bytes(xlsx_bytes)
        records = _parse_ward_totals(wb, year)
        all_records.extend(records)
        print(f"    -> {len(records)} candidate-ward rows")

    return all_records


def fetch_registered_electors() -> list[dict]:
    """Fetch and parse per-ward registered elector counts."""
    all_records: list[dict] = []

    # --- General election voter stats ---
    print("Fetching general election voter statistics...")
    resources = fetch_resources(VOTER_STATS_PACKAGE)
    for resource in resources:
        name = resource.get("name", "")
        year = VOTER_STATS_NAMES.get(name)
        if year is None:
            continue
        fmt = resource.get("format", "").upper()
        print(f"  Downloading {name} ({year})...")
        raw_bytes = download_bytes(resource["url"])
        if fmt == "ZIP":
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
                xlsx_name = next(
                    (n for n in zf.namelist() if n.lower().endswith(".xlsx")), None
                )
                if not xlsx_name:
                    raise ValueError(f"{name}: no XLSX inside ZIP")
                raw_bytes = zf.read(xlsx_name)
        wb = _load_workbook_bytes(raw_bytes, resolve_formulas=True)
        records = _parse_eligible_electors(wb, year)
        all_records.extend(records)
        print(f"    -> {len(records)} ward rows")

    # --- 2023 mayoral by-election voter stats ---
    print("Fetching 2023 by-election voter statistics...")
    by_resources = fetch_resources(BY_ELECTION_VOTER_STATS_PACKAGE)
    for resource in by_resources:
        name = resource.get("name", "")
        year = next(
            (v for k, v in BY_ELECTION_VOTER_STATS_NAMES.items() if k in name), None
        )
        if year is None:
            continue
        print(f"  Downloading {name}...")
        xlsx_bytes = download_bytes(resource["url"])
        wb = _load_workbook_bytes(xlsx_bytes, resolve_formulas=True)
        records = _parse_eligible_electors(wb, year)
        all_records.extend(records)
        print(f"    -> {len(records)} ward rows")

    return all_records


def write_with_sidecar(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)
    sidecar = path.with_suffix(".json")
    sidecar.write_text(
        json.dumps({"fetched_at": datetime.now(timezone.utc).isoformat()}, indent=2),
        encoding="utf-8",
    )
    print(f"  Written: {path} ({len(df)} rows)")


def main() -> None:
    results = fetch_mayoral_results()
    if not results:
        raise RuntimeError("No mayoral results fetched — check package IDs and network")
    df_results = pd.DataFrame(results)
    write_with_sidecar(df_results, OUTPUT_DIR / "mayoral_results.csv")

    electors = fetch_registered_electors()
    if not electors:
        raise RuntimeError(
            "No registered elector data fetched — check package IDs and network"
        )
    df_electors = pd.DataFrame(electors)
    write_with_sidecar(df_electors, OUTPUT_DIR / "registered_electors.csv")

    print("Done.")


if __name__ == "__main__":
    main()
