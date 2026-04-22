#!/usr/bin/env python3
"""Fetch Toronto ward population data from Ward Profiles census data.

Downloads the Ward Profiles (25-Ward Model) XLSX from Toronto Open Data,
extracts total population counts for 2016 and 2021 per ward, and saves a
clean CSV to data/raw/census/.

Outputs:
  data/raw/census/ward_population.csv  -- ward x pop_2016 x pop_2021
  data/raw/census/ward_population.json -- sidecar with fetch timestamp

Run: uv run scripts/fetch_ward_profiles.py
"""

from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import requests

CKAN_BASE = "https://ckan0.cf.opendata.inter.prod-toronto.ca/api/3/action"

WARD_PROFILES_PACKAGE = "6678e1a6-d25f-4dff-b2b7-aa8f042bc2eb"

OUTPUT_DIR = Path("data/raw/census")

# Row indices (0-based) in the ward profiles sheets
_HEADER_ROW_IDX = 17   # [None, "Toronto", "Ward 1", ..., "Ward 25"]
_POP_ROW_IDX = 18      # "Total - Age" row with population totals

_SHEET_2021 = "2021 One Variable"
_SHEET_2016 = "2016 Census One Variable"


# ---------------------------------------------------------------------------
# CKAN helpers
# ---------------------------------------------------------------------------


def fetch_resources(package_id: str) -> list[dict]:
    url = f"{CKAN_BASE}/package_show"
    r = requests.get(url, params={"id": package_id}, timeout=30)
    r.raise_for_status()
    data = r.json()
    if not data["success"]:
        raise RuntimeError(f"CKAN API error for {package_id}: {data.get('error')}")
    return data["result"]["resources"]


def download_bytes(url: str, timeout: int = 120) -> bytes:
    r = requests.get(url, timeout=timeout)
    r.raise_for_status()
    return r.content


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------


def _parse_ward_population(wb: openpyxl.Workbook) -> list[dict]:
    """Extract per-ward population totals for 2016 and 2021.

    Expects:
      Sheet "2021 One Variable": row 17 = column headers, row 18 = population
      Sheet "2016 Census One Variable": same layout

    Returns a list of 25 dicts: {"ward": int, "pop_2016": int, "pop_2021": int}
    """
    expected_headers = [None, "Toronto"] + [f"Ward {i}" for i in range(1, 26)]

    def _read_sheet(sheet_name: str) -> list[int]:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Trim to the first 27 columns (label + Toronto + Ward 1–25); the sheet
        # may have additional empty trailing columns.
        header_row = list(rows[_HEADER_ROW_IDX])[:27]
        if header_row != expected_headers:
            raise ValueError(
                f"Sheet '{sheet_name}' row {_HEADER_ROW_IDX} does not match expected "
                f"headers.\n  Expected: {expected_headers}\n  Got: {header_row}"
            )

        pop_row = list(rows[_POP_ROW_IDX])
        # Columns 2–26 are Ward 1–25 (col 0 = label, col 1 = Toronto city-wide)
        ward_pops = pop_row[2:27]
        result = []
        for i, v in enumerate(ward_pops):
            if v is None or not isinstance(v, (int, float)):
                raise ValueError(
                    f"Sheet '{sheet_name}': unexpected non-numeric population value "
                    f"at ward column {i + 1}: {v!r}"
                )
            result.append(int(v))
        return result

    print(f"  Parsing sheet '{_SHEET_2021}'...")
    pops_2021 = _read_sheet(_SHEET_2021)

    print(f"  Parsing sheet '{_SHEET_2016}'...")
    pops_2016 = _read_sheet(_SHEET_2016)

    if len(pops_2021) != 25 or len(pops_2016) != 25:
        raise RuntimeError(
            f"Expected 25 ward population values, got "
            f"pop_2021={len(pops_2021)}, pop_2016={len(pops_2016)}"
        )

    records = [
        {"ward": ward, "pop_2016": pops_2016[i], "pop_2021": pops_2021[i]}
        for i, ward in enumerate(range(1, 26))
    ]

    return records


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def write_with_sidecar(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)
    sidecar = path.with_suffix(".json")
    sidecar.write_text(
        json.dumps({"fetched_at": datetime.now(timezone.utc).isoformat()}, indent=2),
        encoding="utf-8",
    )
    print(f"  Written: {path} ({len(df)} rows)")


# ---------------------------------------------------------------------------
# Main fetch logic
# ---------------------------------------------------------------------------


def main() -> None:
    print("Fetching Ward Profiles package resources...")
    resources = fetch_resources(WARD_PROFILES_PACKAGE)

    # Find the CensusData XLSX resource
    census_resource = None
    for resource in resources:
        name = resource.get("name", "")
        fmt = resource.get("format", "").upper()
        if "censusdata" in name.lower() and fmt == "XLSX":
            census_resource = resource
            break

    if census_resource is None:
        available = [(r.get("name"), r.get("format")) for r in resources]
        raise RuntimeError(
            f"Could not find CensusData XLSX resource. Available: {available}"
        )

    print(f"  Found resource: {census_resource['name']}")
    print(f"  Downloading from {census_resource['url']} ...")
    xlsx_bytes = download_bytes(census_resource["url"])
    print(f"  Downloaded {len(xlsx_bytes):,} bytes")

    print("  Loading workbook (this may take a moment for large files)...")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    try:
        records = _parse_ward_population(wb)
    finally:
        wb.close()

    # Sanity check population values
    for rec in records:
        for col in ("pop_2016", "pop_2021"):
            val = rec[col]
            if not (50_000 <= val <= 200_000):
                raise RuntimeError(
                    f"Ward {rec['ward']} {col}={val:,} is outside the plausible range "
                    f"50,000–200,000 — check sheet layout"
                )

    df = pd.DataFrame(records)
    write_with_sidecar(df, OUTPUT_DIR / "ward_population.csv")

    print("Done.")


if __name__ == "__main__":
    main()
